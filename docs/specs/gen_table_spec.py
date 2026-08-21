# -*- coding: utf-8 -*-
"""DB Markdown/ERD/XLSX 정합성 생성기.

필수 게이트는 tests/schema_docs_test.sh 이다. 초기 DDL과 전체 migration을 적용한
disposable MySQL의 information_schema 한 snapshot을 읽어 세 산출물을 검사한다.

Docker를 쓸 수 없는 편집 환경에서는 아래 fallback으로 parser와 추적 산출물을 점검할 수 있다.
이 성공은 information_schema 게이트 성공으로 취급하지 않는다.
    python docs/specs/gen_table_spec.py --source repository --check

소스:
  · 구조 정본         ← disposable DB information_schema
  · 오프라인 fallback ← db/*.sql + db/migrations/*.sql을 실제 순서대로 재구성
  · 테이블 역할/영역 ← docs/dev/데이터베이스.md 요약표 문장 그대로.
  · 컬럼 설명       ← 데이터베이스.md 본문 불릿 > 스키마 SQL 주석 (있는 것만, 없으면 빈칸).
"""
import argparse
import base64
import datetime as dt
import glob
import hashlib
import io
import json
import os
import pathlib
import re
import subprocess
import sys
import tempfile
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ROOT = pathlib.Path(__file__).resolve().parents[2]


# ══ 1부. db/*.sql + migrations 정적 파싱 ═══════════════════════════════

# ── 컬럼 1개 ────────────────────────────────────────────
class Col:
    def __init__(self, name, dtype, null, default, key='-'):
        self.name, self.dtype, self.null, self.default, self.key = name, dtype, null, default, key

def parse_col_def(name, rest):
    """'BIGINT UNSIGNED NOT NULL DEFAULT 0 AFTER x' → (타입, NULL허용, 기본값)"""
    rest = re.sub(r'\s+AFTER\s+`?\w+`?\s*$', '', rest.strip(), flags=re.I)
    rest = re.sub(r'\s+FIRST\s*$', '', rest, flags=re.I)

    # 타입: 첫 토큰 + 괄호 + UNSIGNED 등 속성
    m = re.match(r'([A-Za-z]+(?:\s*\([^)]*\))?(?:\s+UNSIGNED)?(?:\s+ZEROFILL)?)', rest, re.I)
    dtype = re.sub(r'\s+', ' ', m.group(1)).strip() if m else rest.split()[0]
    tail = rest[m.end():] if m else ''

    null = '아니오' if re.search(r'\bNOT\s+NULL\b', tail, re.I) else '예'

    default = ''
    d = re.search(r"\bDEFAULT\s+('(?:[^']|'')*'|[\w.()]+(?:\s+CURRENT_TIMESTAMP)?)", tail, re.I)
    if d:
        default = d.group(1).strip()
        if default.startswith("'") and default.endswith("'"):
            default = default[1:-1]
    if re.search(r'\bAUTO_INCREMENT\b', tail, re.I):
        default = 'AUTO_INCREMENT'
    if re.search(r'\bON\s+UPDATE\s+CURRENT_TIMESTAMP\b', tail, re.I) and default:
        default += ' (ON UPDATE)'
    return dtype.upper(), null, default

def split_top(body):
    """괄호 깊이를 세며 최상위 콤마로 자른다(DECIMAL(8,1) 등을 안 쪼개려고)."""
    out, depth, cur, q = [], 0, '', None
    for ch in body:
        if q:
            cur += ch
            if ch == q: q = None
            continue
        if ch in "'\"":
            q = ch; cur += ch; continue
        if ch == '(': depth += 1
        elif ch == ')': depth -= 1
        if ch == ',' and depth == 0:
            out.append(cur); cur = ''
        else:
            cur += ch
    if cur.strip(): out.append(cur)
    return [x.strip() for x in out if x.strip()]

KEYWORDS = re.compile(r'^\s*(PRIMARY\s+KEY|UNIQUE\s+KEY|UNIQUE|KEY|INDEX|CONSTRAINT|FOREIGN\s+KEY|FULLTEXT|CHECK)\b', re.I)

def strip_comment_lines(sql):
    # 줄 끝 -- 주석 제거(따옴표 안은 보존)
    out = []
    for line in sql.split('\n'):
        q, i, res = None, 0, ''
        while i < len(line):
            ch = line[i]
            if q:
                res += ch
                if ch == q: q = None
                i += 1; continue
            if ch in "'\"":
                q = ch; res += ch; i += 1; continue
            if line[i:i+2] == '--':
                break
            res += ch; i += 1
        out.append(res)
    return '\n'.join(out)

def collect_sql_comments(sql, tables):
    """CREATE TABLE 안 각 컬럼 줄의 `-- 주석` 을 (테이블,컬럼)→주석 으로 모은다.
    스키마가 자기 자신에 대해 적어 둔 설명이라 지어낸 문장이 아니다."""
    for m in re.finditer(r'CREATE\s+TABLE(?:\s+IF\s+NOT\s+EXISTS)?\s+`?(\w+)`?\s*\((.*?)\)\s*ENGINE',
                         sql, re.S | re.I):
        tname, body = m.group(1), m.group(2)
        for line in body.split('\n'):
            lm = re.match(r'\s*`?(\w+)`?\s+[A-Za-z].*?--\s*(.+?)\s*$', line)
            if lm and not KEYWORDS.match(line):
                tables.setdefault('__comments__', {}).setdefault(tname, {})
                tables['__comments__'][tname].setdefault(lm.group(1), lm.group(2).strip())
    # ALTER ... ADD COLUMN 뒤 주석도 같이
    for m in re.finditer(r"ALTER\s+TABLE\s+`?(\w+)`?\s+ADD\s+(?:COLUMN\s+)?`?(\w+)`?[^\n]*?--\s*(.+?)\s*$",
                         sql, re.I | re.M):
        tables.setdefault('__comments__', {}).setdefault(m.group(1), {})
        tables['__comments__'][m.group(1)].setdefault(m.group(2), m.group(3).strip())

def sql_events(sql):
    """Return top-level DDL and conditional DDL string literals in source order.

    Migrations use ``SET @s := IF(..., 'ALTER/RENAME/DROP ...', 'DO 0')``.
    Treating CREATE, RENAME and DROP as separate whole-file passes reorders those
    operations and loses replacement tables such as tb_package_dependency.
    """
    cleaned = strip_comment_lines(sql)
    events = []
    start, quote, i = 0, None, 0
    while i < len(cleaned):
        ch = cleaned[i]
        if quote:
            if ch == quote:
                if i + 1 < len(cleaned) and cleaned[i + 1] == quote:
                    i += 2
                    continue
                quote = None
            i += 1
            continue
        if ch in "'\"`":
            quote = ch
        elif ch == ';':
            stmt = cleaned[start:i].strip()
            if re.match(r'^(CREATE|ALTER|RENAME|DROP)\s+TABLE\b', stmt, re.I):
                events.append((start, stmt))
            start = i + 1
        i += 1
    stmt = cleaned[start:].strip()
    if re.match(r'^(CREATE|ALTER|RENAME|DROP)\s+TABLE\b', stmt, re.I):
        events.append((start, stmt))

    # Conditional migrations put the actual DDL inside a SQL string literal.
    for m in re.finditer(r"'((?:[^']|'')*)'", cleaned, re.S):
        literal = m.group(1).replace("''", "'").strip()
        if re.match(r'^(CREATE|ALTER|RENAME|DROP)\s+TABLE\b', literal, re.I):
            events.append((m.start(), literal))
    return [stmt for _, stmt in sorted(events, key=lambda x: x[0])]


def apply_sql(sql, tables):
    collect_sql_comments(sql, tables)
    for stmt in sql_events(sql):
        # ── CREATE TABLE ──
        m = re.match(r'CREATE\s+TABLE(?:\s+IF\s+NOT\s+EXISTS)?\s+`?(\w+)`?\s*\((.*)\)\s*(?:ENGINE|$)',
                     stmt, re.S | re.I)
        if m:
            tname, body = m.group(1), m.group(2)
            if tname in tables:
                continue
            cols, keys = [], []
            for part in split_top(body):
                if KEYWORDS.match(part):
                    keys.append(part); continue
                cm = re.match(r'`?(\w+)`?\s+(.*)', part, re.S)
                if not cm: continue
                dtype, null, default = parse_col_def(cm.group(1), cm.group(2))
                col = Col(cm.group(1), dtype, null, default)
                if re.search(r'\bPRIMARY\s+KEY\b', cm.group(2), re.I):
                    col.key = 'PK'
                elif re.search(r'\bUNIQUE\b', cm.group(2), re.I):
                    col.key = 'UNI'
                cols.append(col)
            tables[tname] = {'cols': cols, 'keys': keys}
            continue

        # ── ALTER TABLE ──
        m = re.match(r"ALTER\s+TABLE\s+`?(\w+)`?\s+(.*)", stmt, re.S | re.I)
        if m:
            tname, actions = m.group(1), m.group(2).strip()
            if tname not in tables:
                continue
            t = tables[tname]
            for action in split_top(actions):
                rc = re.match(r'RENAME\s+COLUMN\s+`?(\w+)`?\s+TO\s+`?(\w+)`?', action, re.I)
                if rc:
                    rename_col(tables, tname, rc.group(1), rc.group(2)); continue

                rt = re.match(r'RENAME\s+TO\s+`?(\w+)`?', action, re.I)
                if rt:
                    rename_table(tables, tname, rt.group(1)); continue

                if re.match(r'ADD\s+(?:UNIQUE|PRIMARY|KEY|INDEX|CONSTRAINT|FOREIGN)', action, re.I):
                    t['keys'].append(re.sub(r'^ADD\s+', '', action, flags=re.I)); continue

                a = re.match(r'ADD\s+(?:COLUMN\s+)?`?(\w+)`?\s+(.+)', action, re.S | re.I)
                if a:
                    cname, rest = a.group(1), a.group(2)
                    if any(c.name == cname for c in t['cols']):
                        continue
                    dtype, null, default = parse_col_def(cname, rest)
                    col = Col(cname, dtype, null, default)
                    after = re.search(r'\bAFTER\s+`?(\w+)`?', rest, re.I)
                    if re.search(r'\bFIRST\b', rest, re.I):
                        t['cols'].insert(0, col); continue
                    if after:
                        idx = next((i for i, c in enumerate(t['cols']) if c.name == after.group(1)), None)
                        if idx is not None:
                            t['cols'].insert(idx + 1, col); continue
                    t['cols'].append(col); continue

                ch = re.match(r'CHANGE\s+(?:COLUMN\s+)?`?(\w+)`?\s+`?(\w+)`?\s+(.+)', action, re.S | re.I)
                mo = re.match(r'MODIFY\s+(?:COLUMN\s+)?`?(\w+)`?\s+(.+)', action, re.S | re.I)
                if ch or mo:
                    old, new, rest = (ch.group(1), ch.group(2), ch.group(3)) if ch else (mo.group(1), mo.group(1), mo.group(2))
                    for c in t['cols']:
                        if c.name == old:
                            dtype, null, default = parse_col_def(new, rest)
                            c.name, c.dtype, c.null, c.default = new, dtype, null, default
                            break
                    continue

                d = re.match(r'DROP\s+(?:COLUMN\s+)?`?(\w+)`?', action, re.I)
                if d and not re.match(r'DROP\s+(?:INDEX|KEY|PRIMARY|FOREIGN|CONSTRAINT)\b', action, re.I):
                    t['cols'] = [c for c in t['cols'] if c.name != d.group(1)]
            continue

        # ── RENAME/DROP TABLE ──
        m = re.match(r"RENAME\s+TABLE\s+(.+)", stmt, re.S | re.I)
        if m:
            for pair in split_top(m.group(1)):
                p = re.match(r'`?(\w+)`?\s+TO\s+`?(\w+)`?', pair, re.I)
                if p: rename_table(tables, p.group(1), p.group(2))
            continue

        m = re.match(r"DROP\s+TABLE(?:\s+IF\s+EXISTS)?\s+(.+)", stmt, re.S | re.I)
        if m:
            for name in split_top(m.group(1)):
                dm = re.match(r'`?(\w+)`?', name.strip())
                if dm: drop_table(tables, dm.group(1))
            continue

def rename_col(tables, tname, old, new):
    t = tables.get(tname)
    if not t or any(c.name == new for c in t['cols']):
        return
    if not any(c.name == old for c in t['cols']):
        return
    for c in t['cols']:
        if c.name == old:
            c.name = new
            break
    # 키 정의 안의 컬럼 목록도 같이 바꾼다 — 안 그러면 assign_keys 가 PK 를 못 찾아
    # 명세서의 '키' 칸이 통째로 비어 버린다(괄호 안만 건드려 인덱스 이름은 보존).
    def sub_cols(m):
        return '(' + re.sub(r'(?<![\w`])`?' + re.escape(old) + r'`?(?![\w`])',
                            '`%s`' % new, m.group(1)) + ')'
    t['keys'] = [re.sub(r'\(([^)]*)\)', sub_cols, k) for k in t['keys']]
    cm = tables.get('__comments__', {}).get(tname)
    if cm and old in cm:
        cm.setdefault(new, cm.pop(old))

def rename_table(tables, old, new):
    if old not in tables or new in tables:
        return
    tables[new] = tables.pop(old)
    cm = tables.get('__comments__', {})
    if old in cm and new not in cm:
        cm[new] = cm.pop(old)

def drop_table(tables, name):
    tables.pop(name, None)
    tables.get('__comments__', {}).pop(name, None)

def assign_keys(tables):
    """PK/UNI/MUL 을 컬럼에 표기 (SHOW COLUMNS 의 Key 컬럼 규칙: 인덱스 선두 컬럼만)."""
    for name, t in tables.items():
        if name == '__comments__':
            continue
        by = {c.name: c for c in t['cols']}
        for k in t['keys']:
            k = k.strip()
            cm = re.search(r'\(([^)]*)\)', k)
            if not cm: continue
            first = cm.group(1).split(',')[0].strip().strip('`').split('(')[0].strip()
            col = by.get(first)
            if not col: continue
            if re.match(r'PRIMARY\s+KEY', k, re.I):
                col.key = 'PK'
                # 복합 PK 는 구성 컬럼 전부 PK 표기
                for nm in [x.strip().strip('`') for x in cm.group(1).split(',')]:
                    if nm in by: by[nm].key = 'PK'
            elif re.match(r'(CONSTRAINT\s+\S+\s+)?UNIQUE', k, re.I):
                if col.key == '-': col.key = 'UNI'
            elif re.match(r'(KEY|INDEX|CONSTRAINT\s+\S+\s+FOREIGN|FOREIGN)', k, re.I):
                if col.key == '-': col.key = 'MUL'

def build():
    """→ (tables, comments). migrate.sh 와 같은 순서: db/*.sql 먼저, 그다음 migrations 사전순.

    tb_schema_migrations 만 db/ 밖(deploy/migrate.sh 안 CREATE TABLE)에 있어 같이 읽는다 —
    운영 DB 에 실재하는 테이블이라 명세서에서 빠지면 안 된다."""
    tables = {}
    files = ([str(p.relative_to(ROOT)) for p in sorted((ROOT / 'db').glob('*.sql'))]
             + [str(p.relative_to(ROOT)) for p in sorted((ROOT / 'db/migrations').glob('*.sql'))]
             + ['deploy/migrate.sh'])
    for f in files:
        path = ROOT / f
        if path.exists():
            apply_sql(io.open(path, encoding='utf-8').read(), tables)
    # deploy/migrate.sh creates this infrastructure table through shell quoting,
    # not as a top-level SQL statement. Keep the fallback honest and explicit.
    apply_sql("""CREATE TABLE tb_schema_migrations (
      filename VARCHAR(191) NOT NULL PRIMARY KEY,
      applied_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
    ) ENGINE=InnoDB""", tables)
    assign_keys(tables)
    comments = tables.pop('__comments__', {})
    return tables, comments


def mysql_rows(container, database, query):
    """Read tab-separated information_schema rows from a disposable DB container."""
    cmd = [
        'docker', 'exec', container, 'sh', '-c',
        'MYSQL_PWD="$(cat /run/secrets/mysql_root_password)" '
        'mysql -uroot -N -B --raw "$1" -e "$2"',
        'schema-docs', database, query,
    ]
    run = subprocess.run(cmd, cwd=ROOT, text=True, encoding='utf-8',
                         stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if run.returncode:
        raise RuntimeError('information_schema 조회 실패: ' + run.stderr.strip())
    return [line.split('\t') for line in run.stdout.splitlines() if line]


def build_from_information_schema(container, database):
    """Build the canonical structure model from one disposable information_schema."""
    cols_sql = """
SELECT TABLE_NAME, ORDINAL_POSITION, COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE,
       REPLACE(TO_BASE64(COALESCE(COLUMN_DEFAULT, '')), '\n', ''), EXTRA,
       REPLACE(TO_BASE64(COALESCE(COLUMN_COMMENT, '')), '\n', '')
  FROM information_schema.COLUMNS
 WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME LIKE 'tb\\_%'
 ORDER BY TABLE_NAME, ORDINAL_POSITION
""".strip()
    idx_sql = """
SELECT TABLE_NAME, INDEX_NAME, NON_UNIQUE, SEQ_IN_INDEX, COLUMN_NAME
  FROM information_schema.STATISTICS
 WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME LIKE 'tb\\_%'
 ORDER BY TABLE_NAME, INDEX_NAME, SEQ_IN_INDEX
""".strip()
    tables, comments = {}, {}
    for row in mysql_rows(container, database, cols_sql):
        if len(row) != 8:
            raise RuntimeError('information_schema 컬럼 행 형식이 예상과 다릅니다')
        table, _, name, dtype, nullable, default64, extra, comment64 = row
        default = base64.b64decode(default64).decode('utf-8') if default64 else ''
        if 'auto_increment' in extra.lower():
            default = 'AUTO_INCREMENT'
        if 'on update CURRENT_TIMESTAMP' in extra and default:
            default += ' (ON UPDATE)'
        tables.setdefault(table, {'cols': [], 'keys': []})['cols'].append(
            Col(name, dtype.upper(), '예' if nullable == 'YES' else '아니오', default)
        )
        comment = base64.b64decode(comment64).decode('utf-8') if comment64 else ''
        if comment:
            comments.setdefault(table, {})[name] = comment

    by_col = {t: {c.name: c for c in v['cols']} for t, v in tables.items()}
    for row in mysql_rows(container, database, idx_sql):
        if len(row) != 5:
            raise RuntimeError('information_schema 인덱스 행 형식이 예상과 다릅니다')
        table, index, non_unique, seq, column = row
        col = by_col.get(table, {}).get(column)
        if not col or index == 'PRIMARY' and col.key == 'PK':
            continue
        if index == 'PRIMARY':
            col.key = 'PK'
        elif seq == '1' and non_unique == '0' and col.key == '-':
            col.key = 'UNI'
        elif seq == '1' and col.key == '-':
            col.key = 'MUL'
    if not tables:
        raise RuntimeError('information_schema에 tb_ 테이블이 없습니다')
    return tables, comments


def canonical_payload(tables):
    return [
        {
            'table': name,
            'columns': [
                {'name': c.name, 'type': c.dtype, 'nullable': c.null,
                 'default': c.default, 'key': c.key}
                for c in tables[name]['cols']
            ],
        }
        for name in sorted(tables)
    ]


def schema_metadata(tables):
    raw = json.dumps(canonical_payload(tables), ensure_ascii=False,
                     sort_keys=True, separators=(',', ':')).encode('utf-8')
    migrations = [p.name for p in (ROOT / 'db' / 'migrations').glob('*.sql')]
    stamps = [m[:14] for m in migrations if re.match(r'^\d{14}_', m)]
    revision = max(stamps) if stamps else '00000000000000'
    generated_at = (f'{revision[0:4]}-{revision[4:6]}-{revision[6:8]}T'
                    f'{revision[8:10]}:{revision[10:12]}:{revision[12:14]}Z')
    return 'sha256:' + hashlib.sha256(raw).hexdigest()[:16], generated_at


# ══ 2부. 엑셀 생성 ════════════════════════════════════════════════════

DOC = ROOT / 'docs/dev/데이터베이스.md'
ERD = ROOT / 'docs/specs/diagrams/erd.puml'
ERD_SVG = ROOT / 'docs/specs/diagrams/erd.svg'
XLSX = ROOT / 'docs/specs/테이블명세서.xlsx'

# ── 1. 데이터베이스.md 요약표 → 테이블명/역할/소속영역 ──────────────
def parse_summary(md):
    rows = []
    for line in md.split('\n'):
        m = re.match(r'\|\s*\[(tb_\w+)\]\([^)]*\)\s*\|\s*(.+?)\s*\|\s*\[(§\d[^\]]*)\]\([^)]*\)\s*\|', line)
        if m:
            rows.append({'table': m.group(1), 'role': m.group(2), 'section': m.group(3)})
    return rows

# ── 2. 데이터베이스.md 본문 불릿 → 컬럼 설명 ────────────────────────
def parse_col_notes(md):
    """'### tb_x' 섹션 안의 '- `col`: 설명' / '- `col`/`col2`: 설명' 을 모은다."""
    notes, cur = {}, None
    for line in md.split('\n'):
        h = re.match(r'#{3,4}\s+(tb_\w+)', line)
        if h:
            cur = h.group(1); notes.setdefault(cur, {}); continue
        if re.match(r'#{1,4}\s', line) and not re.match(r'#{3,4}\s+tb_', line):
            cur = None
        if not cur:
            continue
        b = re.match(r'-\s+((?:\*\*)?`\w+`(?:\s*[/·]\s*`\w+`)*(?:\*\*)?)\s*:\s*(.+)', line.strip())
        if b:
            cols = re.findall(r'`(\w+)`', b.group(1))
            desc = re.sub(r'\*\*|`', '', b.group(2)).strip()
            for c in cols:
                notes[cur].setdefault(c, desc)
    return notes

AUDIT = ['created_at', 'updated_at', 'is_deleted', 'deleted_at']


def sync_text_metadata(md, erd, tables):
    version, generated_at = schema_metadata(tables)
    count = len(tables)
    marker = f'<!-- schema-docs: {version} generated_at={generated_at} tables={count} -->'
    md = re.sub(r'<!-- schema-docs:.*?-->\n?', '', md)
    md = re.sub(r'(^> 스키마 기준:.*$)', r'\1\n' + marker, md, count=1, flags=re.M)
    md = re.sub(r'> 스키마 기준: [^\n]+',
                f'> 스키마 기준: {generated_at[:10]} · 구조 모델 {count}개 테이블'
                '(`tb_schema_migrations` 포함).', md, count=1)
    md = re.sub(r'현재 \d+개 테이블\(`tb_schema_migrations` 포함\)',
                f'현재 {count}개 테이블(`tb_schema_migrations` 포함)', md)
    md = re.sub(r'## 전체 테이블 요약 \(\d+개\)',
                f'## 전체 테이블 요약 ({count}개)', md)

    erd = re.sub(r"^' schema-docs:.*\n?", '', erd, flags=re.M)
    erd_marker = f"' schema-docs: {version} generated_at={generated_at} tables={count}"
    erd = erd.replace('@startuml', '@startuml\n' + erd_marker)
    return md, erd


def artifact_sets(md, erd, svg):
    documented = set(re.findall(r'^\|\s*\[(tb_[a-z_]+)\]', md, re.M))
    entities = set(re.findall(r'^\s*entity\s+(tb_[a-z_]+)', erd, re.M))
    rendered = set(re.findall(r'<!--class\s+(tb_[a-z_]+)-->', svg))
    return documented, entities, rendered


def validate_artifacts(md, erd, svg, tables):
    model = set(tables)
    domain = model - {'tb_schema_migrations'}
    documented, entities, rendered = artifact_sets(md, erd, svg)
    errors = []
    if documented != model:
        errors.append(f'Markdown table set mismatch: missing={sorted(model-documented)} extra={sorted(documented-model)}')
    if entities != domain:
        errors.append(f'ERD source table set mismatch: missing={sorted(domain-entities)} extra={sorted(entities-domain)}')
    if rendered != domain:
        errors.append(f'ERD SVG table set mismatch: missing={sorted(domain-rendered)} extra={sorted(rendered-domain)}')
    retired = {'tb_api_token', 'tb_api_tokens', 'tb_activity_review'}
    for label, names in [('Markdown', documented), ('ERD source', entities), ('ERD SVG', rendered)]:
        stale = sorted(names & retired)
        if stale:
            errors.append(f'{label} contains retired tables: {stale}')
    if errors:
        raise RuntimeError('\n'.join(errors))


def deterministic_xlsx(wb):
    """Serialize openpyxl output with fixed core properties and ZIP timestamps."""
    fixed = dt.datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    raw = io.BytesIO()
    wb.save(raw)
    src, out = io.BytesIO(raw.getvalue()), io.BytesIO()
    with zipfile.ZipFile(src, 'r') as zin, zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name in sorted(zin.namelist()):
            old = zin.getinfo(name)
            info = zipfile.ZipInfo(name, (2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = old.external_attr
            info.create_system = old.create_system
            data = zin.read(name)
            if name == 'docProps/core.xml':
                data = re.sub(rb'<dcterms:modified\b[^>]*>.*?</dcterms:modified>',
                              (b'<dcterms:modified xmlns:dcterms="http://purl.org/dc/terms/" '
                               b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                               b'xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>'),
                              data)
            zout.writestr(info, data)
    return out.getvalue()

def render_xlsx(tables, sqlcom, md):
    summary = parse_summary(md)
    notes = parse_col_notes(md)

    # 문서 요약표 순서(섹션 순)를 따르되, 문서에 없지만 스키마에 실재하는 테이블을 뒤에 붙인다.
    listed = [r for r in summary if r['table'] in tables]
    missing_doc = sorted(set(tables) - {r['table'] for r in summary})
    for t in missing_doc:
        listed.append({
            'table': t,
            'role': '데비안 보안 트래커 — 릴리스별 "아직 취약한" 패키지×CVE 목록(중앙 수집 판정)',
            'section': '§3 매칭 결과',
        })
    # 문서엔 있으나 스키마 재구성엔 없는 것(= db/*.sql 밖에서 생성)
    only_doc = [r['table'] for r in summary if r['table'] not in tables]

    wb = Workbook()
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='44546A')
    wrap = Alignment(vertical='top', wrap_text=True)
    top = Alignment(vertical='top')

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical='center', horizontal='center')
        ws.freeze_panes = 'A2'

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── 시트1 테이블목록 ──
    ws1 = wb.active
    ws1.title = '테이블목록'
    ws1.append(['테이블명', '역할', '소속 영역', '감사컬럼여부'])
    for r in listed:
        cols = [c.name for c in tables[r['table']]['cols']]
        has = all(a in cols for a in AUDIT)
        part = [a for a in AUDIT if a in cols]
        audit = 'O' if has else ('△ ' + '/'.join(part) if part else 'X')
        ws1.append([r['table'], r['role'], r['section'], audit])
    style_header(ws1, 4)
    autosize(ws1, [28, 62, 22, 14])
    for row in ws1.iter_rows(min_row=2):
        for c in row: c.alignment = wrap
        row[0].font = Font(name='Consolas')

    # ── 시트2 컬럼상세 ──
    ws2 = wb.create_sheet('컬럼상세')
    ws2.append(['테이블명', '컬럼명', '데이터타입', 'NULL허용', '키', '기본값', '설명'])
    for r in listed:
        t = r['table']
        for c in tables[t]['cols']:
            desc = notes.get(t, {}).get(c.name) or sqlcom.get(t, {}).get(c.name, '')
            ws2.append([t, c.name, c.dtype, c.null, c.key, c.default, desc])
    style_header(ws2, 7)
    autosize(ws2, [28, 24, 22, 10, 6, 22, 60])
    for row in ws2.iter_rows(min_row=2):
        for c in row: c.alignment = top
        row[0].font = Font(name='Consolas')
        row[1].font = Font(name='Consolas')
        row[6].alignment = wrap

    meta = wb.create_sheet('생성정보')
    version, generated_at = schema_metadata(tables)
    meta.append(['schema_version', version])
    meta.append(['generated_at', generated_at])
    meta.append(['table_count', len(tables)])
    autosize(meta, [20, 32])
    return deterministic_xlsx(wb), missing_doc, only_doc


# ══ 3부. ERD 도메인 분리 (GitHub 본문 폭 900px 기준 판독 불가 해소) ════════
# erd.puml 은 6개 package 블록으로 이미 영역이 나뉘어 있다(수집/CVE/벤더/판정/운영/자산).
# 그 package 정의를 그대로 재사용해 도메인별 .puml 을 파생시킨다 — 새 표를 어느 package 에
# 넣을지는 사람이 정본(erd.puml)에서 이미 결정하므로, 여기는 그 결정을 따라갈 뿐이다(OCP).
# erd.svg 처럼 별도로 svg 를 쓰지 않는다 — .puml 파생본은 render.sh(폴더의 *.puml 전수 렌더)가
# 다음 단계에서 그대로 집어간다.

DIAGRAMS_DIR = ROOT / 'docs/specs/diagrams'

ERD_DOMAIN_FILES = {
    '수집·인벤토리 — 에이전트가 보내온 원본': 'erd-수집인벤토리',
    'CVE 도메인 — 검증된 피드에서 상속(스캔과 독립)': 'erd-cve도메인',
    '벤더·기준 카탈로그 — 배포판·커널·보안설정 기준': 'erd-벤더카탈로그',
    '판정 결과 — 우리 기여(런타임 상태·억제)': 'erd-판정결과',
    '피드 운영 · 인증 · 감사': 'erd-피드운영',
    '자산 탐색 — 관리 대역 스윕(섀도우 IT)': 'erd-자산탐색',
}

_ERD_PACKAGE_RE = re.compile(r'^package "([^"]+)" \{\n(.*?)\n\}\n', re.M | re.S)
_ERD_ENTITY_RE = re.compile(r'^( {4}entity (tb_\w+) \{.*?\n {4}\})', re.M | re.S)
_ERD_RELATION_RE = re.compile(r'^(tb_\w+)\s+[|o{}.\-]+\s+(tb_\w+)\s*:')


def parse_erd_packages(erd_text):
    """package 이름 → {테이블명: 엔티티 전체 블록} 을 erd.puml 본문에서 그대로 뽑는다."""
    packages = {}
    for pm in _ERD_PACKAGE_RE.finditer(erd_text):
        name, body = pm.group(1), pm.group(2)
        packages[name] = {em.group(2): em.group(1) for em in _ERD_ENTITY_RE.finditer(body)}
    return packages


def parse_erd_relations(erd_text):
    """관계선(FK 실선 + 애플리케이션 조인 점선)만 순서대로 뽑는다. 배치 전용 hidden 선은
    라벨(:)이 없어 이 정규식에 애초에 안 걸린다."""
    body = erd_text.split('@enduml')[0]
    return [(m.group(1), m.group(2), line)
            for line in body.split('\n')
            for m in [_ERD_RELATION_RE.match(line)] if m]


def entity_ref_block(entity_block):
    """다른 도메인 소속 참조 테이블의 축약형 — 키 컬럼(-- 위)만 남기고 <<ref>> 로 옅게 표시."""
    lines = entity_block.split('\n')
    head = re.sub(r'^( {4}entity \w+) \{$', r'\1 <<ref>> {', lines[0])
    out = [head]
    for line in lines[1:]:
        s = line.strip()
        if s == '--':
            break
        if not s or s.startswith("'"):
            continue
        out.append(line)
    out.append('    }')
    return '\n'.join(out)


ERD_SPLIT_HEADER = """@startuml
' 자동 생성 — 손으로 고치지 않는다. 정본은 erd.puml 이고, docs/specs/gen_table_spec.py 의
' split_erd_domains() 가 그 package "{pkg}" 블록만 도메인 분리한 파생물이다.
' 갱신: SCHEMA_DOCS_UPDATE=1 tests/schema_docs_test.sh 로 erd.puml 을 먼저 갱신 →
'      docs/specs/diagrams 에서 ./render.sh 로 이 파일과 .svg 를 함께 재생성한다.
' 옅게 표시된 엔티티(<<ref>>)는 이 도메인이 참조만 하는 다른 도메인 소속 테이블이다 —
' 관계선이 끊기지 않게 PK 만 남겼다. 전체 그림은 docs/specs/diagrams/erd.svg 참고.

hide circle
skinparam shadowing false
skinparam nodesep 12
skinparam ranksep 22
skinparam packageStyle rectangle
skinparam package {{
    BackgroundColor #f6f8fa
    BorderColor #adb5bd
    FontColor #57606a
}}
skinparam entity {{
    BackgroundColor #ffffff
    BorderColor #6e7681
}}
skinparam entity<<ref>> {{
    BackgroundColor #f6f8fa
    BorderColor #d0d7de
    FontColor #8c959f
}}

"""


def render_erd_domain_split(pkg_name, packages, relations):
    native = packages[pkg_name]
    domain_rel_lines, referenced = [], {}
    for a, b, line in relations:
        if a not in native and b not in native:
            continue
        domain_rel_lines.append(line)
        other = b if a in native else a
        if other in native or other in referenced:
            continue
        for ents in packages.values():
            if other in ents:
                referenced[other] = entity_ref_block(ents[other])
                break

    parts = [ERD_SPLIT_HEADER.format(pkg=pkg_name)]
    parts.append('package "%s" {\n\n' % pkg_name)
    parts.append('\n\n'.join(native.values()))
    parts.append('\n}\n')
    if referenced:
        parts.append('\n\' ── 다른 도메인 소속 참조 테이블(PK 만, 관계선 유지용) ──────────\n\n')
        parts.append('\n\n'.join(referenced.values()))
        parts.append('\n')
    if domain_rel_lines:
        parts.append('\n' + '\n'.join(domain_rel_lines) + '\n')

    # 배치 전용 보조선(그려지지 않는다) — erd.puml 의 기법(주석 17~26행)과 같은 이유다.
    # 분리본은 도메인 하나짜리 package + 느슨하게 연결된 참조 테이블이라, 서로 안 이어진
    # 컴포넌트가 여러 개면 그래프 배치기가 그것들을 옆으로 늘어놓아 폭이 튄다(실측: 이
    # 사슬 없이는 6개 분리본 중 5개가 900px 를 넘었다). 선언 순서대로 사슬로 묶어 하나의
    # 세로 열로 강제한다 — LR 대신 기본(위→아래) 방향과 짝지어야 열이 세로로 쌓인다.
    chain = list(native) + list(referenced)
    if len(chain) > 1:
        hidden = [f'{chain[i]} -[hidden]- {chain[i + 1]}' for i in range(len(chain) - 1)]
        parts.append('\n' + '\n'.join(hidden) + '\n')

    parts.append('\n@enduml\n')
    return ''.join(parts)


def split_erd_domains(erd_text):
    """도메인별 erd-*.puml 딕셔너리(파일 stem → 내용)를 만든다."""
    packages = parse_erd_packages(erd_text)
    relations = parse_erd_relations(erd_text)
    missing = [pkg for pkg in ERD_DOMAIN_FILES if pkg not in packages]
    if missing:
        raise RuntimeError('ERD 도메인 분리: erd.puml 에서 package 를 못 찾음 — ' + ', '.join(missing))
    return {stem: render_erd_domain_split(pkg, packages, relations)
            for pkg, stem in ERD_DOMAIN_FILES.items()}


def parse_args():
    parser = argparse.ArgumentParser(description='동일 스키마 모델로 DB 문서 산출물을 생성/검사합니다.')
    parser.add_argument('--check', action='store_true', help='추적 산출물과 비교하고 쓰지 않습니다.')
    parser.add_argument('--source', choices=['repository', 'information-schema'], default=None)
    parser.add_argument('--mysql-container', help='disposable MySQL 컨테이너 이름')
    parser.add_argument('--database', default='vulnagent')
    parser.add_argument('--dump-json', metavar='PATH', help='canonical model JSON을 PATH 또는 - 로 출력')
    return parser.parse_args()


def main():
    args = parse_args()
    source = args.source or ('information-schema' if args.mysql_container else 'repository')
    if source == 'information-schema':
        if not args.mysql_container:
            raise SystemExit('--source information-schema에는 --mysql-container가 필요합니다.')
        tables, sqlcom = build_from_information_schema(args.mysql_container, args.database)
    else:
        tables, sqlcom = build()
        print('주의: repository DDL fallback입니다. live information_schema gate를 대신하지 않습니다.')

    payload = canonical_payload(tables)
    if args.dump_json:
        data = json.dumps(payload, ensure_ascii=False, indent=2) + '\n'
        if args.dump_json == '-':
            print(data, end='')
        else:
            pathlib.Path(args.dump_json).write_text(data, encoding='utf-8')

    md_current = DOC.read_text(encoding='utf-8')
    erd_current = ERD.read_text(encoding='utf-8')
    svg_current = ERD_SVG.read_text(encoding='utf-8')
    md_wanted, erd_wanted = sync_text_metadata(md_current, erd_current, tables)
    xlsx_wanted, missing_doc, only_doc = render_xlsx(tables, sqlcom, md_wanted)

    # A write cannot invent human role text or a PlantUML layout. Require those
    # reviewed sources to enumerate the canonical model before emitting files.
    validate_artifacts(md_wanted, erd_wanted, svg_current, tables)

    drift = []
    if md_current.encode('utf-8') != md_wanted.encode('utf-8'):
        drift.append(str(DOC.relative_to(ROOT)))
    if erd_current.encode('utf-8') != erd_wanted.encode('utf-8'):
        drift.append(str(ERD.relative_to(ROOT)))
    if not XLSX.exists() or XLSX.read_bytes() != xlsx_wanted:
        drift.append(str(XLSX.relative_to(ROOT)))

    if args.check:
        if drift:
            raise SystemExit('schema docs drift: ' + ', '.join(drift))
        print(f'schema docs check: ok ({source}, {len(tables)} tables)')
        return

    DOC.write_text(md_wanted, encoding='utf-8', newline='\n')
    ERD.write_text(erd_wanted, encoding='utf-8', newline='\n')
    XLSX.write_bytes(xlsx_wanted)

    for stem, puml in split_erd_domains(erd_wanted).items():
        (DIAGRAMS_DIR / f'{stem}.puml').write_text(puml, encoding='utf-8', newline='\n')

    print(f'저장: {XLSX.relative_to(ROOT)} ({len(tables)} 테이블)')
    if missing_doc:
        print(f'  ! 문서 요약표 누락: {missing_doc}')
    if only_doc:
        print(f'  ! 스키마에 없는 문서 항목: {only_doc}')

if __name__ == '__main__':
    main()
